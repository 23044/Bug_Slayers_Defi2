from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth import get_user_model
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.views import View
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from .models import Word, WordVote, Comment, Challenge, UserContribution, UserProfile, ImportedDocument
from .forms import WordForm, CommentForm
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

User = get_user_model()

class HomeView(TemplateView):
    template_name = 'home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['challenge'] = Challenge.objects.filter(is_active=True).first()
        context['recent_words'] = Word.objects.filter(status='approved').order_by('-created_at')[:5]
        context['total_words'] = Word.objects.count()
        context['total_contributors'] = UserProfile.objects.count()
        return context

class DictionaryView(LoginRequiredMixin, TemplateView):
    template_name = 'dictionary/dictionary.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Liste complète des mots
        context['words'] = Word.objects.all().order_by('-created_at')
        
        # Listes filtrées par statut
        context['pending_words'] = Word.objects.filter(status='pending').order_by('-created_at')
        context['review_words'] = Word.objects.filter(status='review').order_by('-created_at')
        context['approved_words'] = Word.objects.filter(status='approved').order_by('-created_at')
        
        # Statistiques pour le challenge
        total_root_words = Word.objects.filter(
            status='approved',
            word_type='root'
        ).count()
        context['challenge_progress'] = min(int((total_root_words / 1000) * 100), 100)
        context['challenge_completed_words'] = total_root_words
        
        # Top contributeurs
        context['top_contributors'] = User.objects.annotate(
            points=Sum('profile__points'),
            words_contributed=Count('contributed_words')
        ).order_by('-points')[:5]
        
        return context

@login_required
def submit_word(request):
    if request.method == 'POST':
        word = request.POST.get('word')
        definition = request.POST.get('definition')
        example = request.POST.get('example')
        word_type = request.POST.get('word_type', 'derived')
        
        if word and definition:
            # Vérifier si le mot existe déjà
            if Word.objects.filter(word=word).exists():
                messages.error(request, "Ce mot existe déjà dans le dictionnaire.")
                return redirect('dictionary:home')
            
            # Créer le nouveau mot
            new_word = Word.objects.create(
                word=word,
                definition=definition,
                examples=[example] if example else [],
                word_type=word_type,
                contributor=request.user,
                status='pending'
            )
            
            # Mettre à jour les statistiques du contributeur
            profile = request.user.profile
            profile.words_contributed += 1
            profile.points += 10  # Points pour la soumission
            profile.save()
            
            messages.success(request, "Votre mot a été soumis avec succès et est en attente de modération.")
        else:
            messages.error(request, "Le mot et la définition sont requis.")
            
    return redirect('dictionary:home')

@login_required
def approve_word(request, word_id):
    if not request.user.is_staff:
        messages.error(request, "Vous n'avez pas la permission d'approuver des mots.")
        return JsonResponse({'error': 'Permission denied'}, status=403)
        
    word = get_object_or_404(Word, id=word_id)
    
    # Ne pas approuver un mot déjà approuvé
    if word.status == 'approved':
        return JsonResponse({'error': 'Ce mot est déjà approuvé'}, status=400)
    
    # Approuver le mot
    word.status = 'approved'
    word.save()
    
    # Mettre à jour les statistiques du contributeur
    if word.contributor:
        profile = word.contributor.profile
        profile.words_contributed += 1
        profile.words_approved += 1
        
        # Points bonus selon le type de mot
        if word.word_type == 'root':
            points_bonus = 100  # Bonus pour les mots racines
        else:
            points_bonus = 50   # Points standard pour les mots dérivés
            
        profile.points += points_bonus
        
        # Mettre à jour le niveau
        profile.level = (profile.points // 100) + 1
        
        profile.save()
        
        # Envoyer une notification au contributeur
        messages.success(request, f"Le mot '{word.word}' a été approuvé et ajouté au dictionnaire.")
        
    return JsonResponse({
        'success': True,
        'message': 'Mot approuvé avec succès',
        'points_awarded': points_bonus
    })

@login_required
def reject_word(request, word_id):
    if not request.user.is_staff:
        messages.error(request, "Vous n'avez pas la permission de rejeter des mots.")
        return JsonResponse({'error': 'Permission denied'}, status=403)
        
    word = get_object_or_404(Word, id=word_id)
    
    # Ne pas rejeter un mot déjà rejeté
    if word.status == 'rejected':
        return JsonResponse({'error': 'Ce mot est déjà rejeté'}, status=400)
    
    # Rejeter le mot
    word.status = 'rejected'
    word.save()
    
    messages.warning(request, f"Le mot '{word.word}' a été rejeté.")
    return JsonResponse({'success': True, 'message': 'Mot rejeté'})

@login_required
def add_comment(request, word_id):
    if request.method == 'POST':
        word = get_object_or_404(Word, id=word_id)
        data = json.loads(request.body)
        
        comment = Comment.objects.create(
            word=word,
            user=request.user,
            text=data.get('text')
        )
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def vote_word(request, word_id):
    word = get_object_or_404(Word, id=word_id)
    
    # Vérifier si l'utilisateur a déjà voté
    if WordVote.objects.filter(user=request.user, word=word).exists():
        return JsonResponse({'error': 'Already voted'}, status=400)
    
    # Créer le vote
    WordVote.objects.create(user=request.user, word=word)
    word.vote_count = word.votes.count()
    word.save()
    
    # Mettre à jour les points du contributeur
    if word.contributor:
        profile = word.contributor.profile
        profile.points += 2  # Points pour chaque vote reçu
        profile.save()
    
    return JsonResponse({'success': True, 'votes': word.vote_count})

class ChallengeView(DetailView):
    model = Challenge
    template_name = 'dictionary/challenge.html'
    context_object_name = 'challenge'

    def get_object(self):
        return Challenge.objects.filter(is_active=True).first()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['top_contributors'] = UserProfile.objects.select_related('user').order_by('-points')[:10]
        context['recent_words'] = Word.objects.filter(
            word_type='root',
            status='approved'
        ).order_by('-created_at')[:5]
        return context

class ModerateWordsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Word
    template_name = 'dictionary/moderate_words.html'
    context_object_name = 'words'
    paginate_by = 20

    def test_func(self):
        return self.request.user.is_staff

    def handle_no_permission(self):
        messages.error(self.request, "Vous n'avez pas la permission d'accéder à la modération.")
        return redirect('home')

    def get_queryset(self):
        queryset = Word.objects.select_related('contributor').order_by('-created_at')
        
        # Filtrer par statut
        status = self.request.GET.get('status', 'pending')
        if status:
            queryset = queryset.filter(status=status)
            
        # Filtrer par type de mot
        word_type = self.request.GET.get('word_type')
        if word_type:
            queryset = queryset.filter(word_type=word_type)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pending_count'] = Word.objects.filter(status='pending').count()
        context['approved_count'] = Word.objects.filter(status='approved').count()
        context['rejected_count'] = Word.objects.filter(status='rejected').count()
        return context

class ManageChallengeView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        challenges = Challenge.objects.all()
        return render(request, 'dictionary/manage_challenge.html', {
            'challenges': challenges
        })

    def post(self, request):
        action = request.POST.get('action')
        
        if action == 'create':
            Challenge.objects.create(
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                target_words=int(request.POST.get('target_words', 1000)),
                start_date=timezone.now()
            )
            messages.success(request, "Challenge créé avec succès!")
            
        elif action == 'update':
            challenge_id = request.POST.get('challenge_id')
            challenge = get_object_or_404(Challenge, id=challenge_id)
            challenge.title = request.POST.get('title')
            challenge.description = request.POST.get('description')
            challenge.target_words = int(request.POST.get('target_words', 1000))
            challenge.save()
            messages.success(request, "Challenge mis à jour avec succès!")
            
        return redirect('dictionary:manage_challenge')

class ImportDocumentView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ImportedDocument
    template_name = 'dictionary/import_document.html'
    fields = ['title', 'file']
    success_url = '/dictionary/import/'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        form.instance.uploaded_by = self.request.user
        response = super().form_valid(form)
        
        # Lancer le traitement du document en arrière-plan
        # TODO: Implémenter la tâche Celery pour le traitement
        
        messages.success(self.request, "Document importé avec succès! Le traitement va commencer.")
        return response

def get_statistics(request):
    total_words = Word.objects.count()
    approved_words = Word.objects.filter(status='approved').count()
    total_contributors = UserProfile.objects.count()
    new_words_this_week = Word.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=7)
    ).count()
    
    return JsonResponse({
        'total_words': total_words,
        'approval_rate': (approved_words / total_words * 100) if total_words > 0 else 0,
        'total_contributors': total_contributors,
        'new_words_this_week': new_words_this_week
    })

def contribution_guide(request):
    return render(request, 'dictionary/contribution_guide.html')

def transcription_rules(request):
    return render(request, 'dictionary/transcription_rules.html')

def contributors_list(request):
    # Get all contributors ordered by points
    contributors = UserProfile.objects.select_related('user').order_by('-points')
    
    # Calculate total statistics
    total_contributions = Word.objects.count()
    validated_words = Word.objects.filter(status='approved').count()
    
    # Pagination
    paginator = Paginator(contributors, 20)  # 20 contributors per page
    page = request.GET.get('page')
    contributors = paginator.get_page(page)
    
    context = {
        'contributors': contributors,
        'total_contributions': total_contributions,
        'validated_words': validated_words,
        'is_paginated': paginator.num_pages > 1,
        'page_obj': contributors,
    }
    
    return render(request, 'dictionary/contributors.html', context)

@login_required
def import_document(request):
    if request.method == 'POST':
        # Logique d'importation de document à implémenter
        messages.info(request, "Fonctionnalité en cours de développement")
        return redirect('dictionary:dictionary')
    return render(request, 'dictionary/import_document.html')

def download_template(request):
    """
    Génère et retourne un modèle Excel pour l'importation de mots.
    """
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle Import"

    # Définir les en-têtes
    headers = [
        'Mot (Hassaniya)',
        'Définition (Français)',
        'Type (racine/dérivé)',
        'Exemples (optionnel)',
        'Variante dialectale (optionnel)'
    ]

    # Style pour les en-têtes
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')

    # Ajouter les en-têtes avec style
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Exemple de données
    example_data = [
        ['kteb', 'écrire', 'racine', 'kteb el-bra (il a écrit une lettre)', 'standard'],
        ['mektûb', 'écrit, lettre', 'dérivé', 'el-mektûb wuṣul (la lettre est arrivée)', 'standard']
    ]

    # Ajouter les exemples
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Sauvegarder dans un buffer
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Préparer la réponse HTTP
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=modele_import_hassaniya.xlsx'

    return response 